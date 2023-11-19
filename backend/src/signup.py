'''
Sign Up
Filename: signup.py

Author: Jacqueline
Created: 28/8/2023
Description: Contains all functions related to signing up.

'''

import pandas as pd
import os
from openpyxl import load_workbook, Workbook
from google.oauth2 import service_account
from googleapiclient.discovery import build
from database import database
from error import InputError
from customer_email import send_email
from flask import send_from_directory

SHEETS_URL = "https://docs.google.com/spreadsheets/d/1q-l5Kt--AAagoRzwOMsO8xJCYRFmcqvE5fR9ahUnm_8/edit#gid=0"
SHEET_ID = "1q-l5Kt--AAagoRzwOMsO8xJCYRFmcqvE5fR9ahUnm_8"

def signup(name: str, email: str)->dict:
    store = database.get()
    users = store['users']

    # Error handling
    if name == None or email == None:
        raise InputError("Please enter a name and an email address")
    emails = [user["email"] for user in users.values()]
    if email in emails:
        raise InputError("Email has already been used")

    # Save email and name
    customer_id = len(users)
    users[customer_id] = { "name": name, "email": email }

    # Set store:
    store['users'] = users
    database.set(store)

    send_email(customer_id)
    update_google_sheets(customer_id)
    return {}

def update_google_sheets(customer_id):
    line_num = int(customer_id) + 2
    store = database.get()
    users = store['users']
    user = users[customer_id]

    customer_details = [user["name"], user["email"]]
    excel_range = "Customers!" + "A" + str(line_num) + ":B" + str(line_num)

    credentials = service_account.Credentials.from_service_account_file("key.json", scopes=["https://www.googleapis.com/auth/spreadsheets"])
    service = build("sheets", "v4", credentials=credentials)
    request = service.spreadsheets().values().update(spreadsheetId=SHEET_ID, range=excel_range, body={"values": [customer_details]}, valueInputOption="RAW")
    request.execute()

def export_to_excel():
    store = database.get()
    users = store['users']

    # Check for excel file, if doesn't exist create it
    output_directory = "src"
    file_name = "Customer Emails.xlsx"
    excel_name = os.path.join(output_directory, file_name)
    s_name = "Current Customers"
<<<<<<< HEAD
    if not os.path.exists(excel_name):
        Wb = Workbook()
        Wb.save(excel_name)
        # Create framework of excel
        df = pd.DataFrame(columns=["Name", "Email"])
        df.to_excel(excel_name, sheet_name=s_name)

    # Create data frame for customer
    data = {'Name': user["name"], 'Email': user["email"]}
    # df = pd.DataFrame(data, index=[0])

    # Append to excel
    existing_sheet = pd.read_excel(excel_name)
    updated_sheet = existing_sheet._append(data, ignore_index=True)
    book = load_workbook(excel_name)
    writer = pd.ExcelWriter(excel_name, engine='openpyxl')
    writer.book = book

    updated_sheet.to_excel(writer, index=False, header=False, sheet_name=s_name)
=======
    # Remove old one
    if os.path.exists(excel_name):
        os.remove(excel_name)
>>>>>>> 83b3b374f83cebe5cf66beaa0231292ebb27b341

    # Create new excel
    Wb = Workbook()
    Wb.save(excel_name)
    # Create framework of excel
    data = [[user["name"], user["email"]] for user in users.values()]
    df = pd.DataFrame(data, columns=["Name", "Email"])
    with pd.ExcelWriter(excel_name) as writer:
        df.to_excel(writer, sheet_name=s_name)
    writer.save()
    writer.close()

    return { "Excel": file_name }